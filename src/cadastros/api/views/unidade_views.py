import io
import re
import secrets
import string
import unicodedata

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ...models import Unidade
from ..serializers import (
    UnidadeCreateBulkSerializer,
    UnidadeListSerializer,
    UnidadeSerializer,
)

User = get_user_model()


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def unidade_list_view(request):
    """
    Lista todas as unidades com paginação e busca.
    - Síndicos e Admin: veem todas as unidades do condomínio
    - Portaria: vê todas as unidades do condomínio
    - Moradores: veem apenas suas próprias unidades
    """
    try:
        user = request.user

        # Busca
        search = request.GET.get("search", "")
        # Relação com usuários é M2M (User.unidades, related_name='moradores')
        unidades = Unidade.objects.prefetch_related("moradores").all()

        # Controle de acesso por grupo
        is_sindico = user.groups.filter(name="Síndicos").exists()
        is_portaria = user.groups.filter(name="Portaria").exists()
        is_morador = user.groups.filter(name="Moradores").exists()

        # Filtrar por condomínio: apenas unidades criadas por usuários do mesmo condomínio
        if hasattr(user, "condominio_id") and user.condominio_id:
            unidades = unidades.filter(
                created_by__condominio_id=user.condominio_id
            )

        if is_morador and not (user.is_staff or is_sindico or is_portaria):
            # Moradores veem apenas suas próprias unidades
            unidades = unidades.filter(moradores=user)
        elif not (user.is_staff or is_sindico or is_portaria):
            # Usuários sem permissão não veem nada
            unidades = Unidade.objects.none()

        if search:
            unidades = unidades.filter(
                Q(numero__icontains=search)
                | Q(bloco__icontains=search)
                | Q(moradores__first_name__icontains=search)
                | Q(moradores__last_name__icontains=search)
            ).distinct()

        # Filtro de status
        is_active = request.GET.get("is_active", None)
        if is_active is not None:
            unidades = unidades.filter(is_active=is_active.lower() == "true")

        # Ordenação
        unidades = unidades.order_by("bloco", "numero")

        # Paginação
        page = int(request.GET.get("page", 1))
        paginator = Paginator(unidades, 12)
        page_obj = paginator.get_page(page)

        serializer = UnidadeListSerializer(page_obj.object_list, many=True)

        return Response(
            {
                "results": serializer.data,
                "count": paginator.count,
                "num_pages": paginator.num_pages,
                "current_page": page_obj.number,
                "has_next": page_obj.has_next(),
                "has_previous": page_obj.has_previous(),
            }
        )

    except Exception as e:
        return Response(
            {"error": f"Erro ao listar unidades: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def unidade_create_view(request):
    """
    Cria uma nova unidade.
    Apenas Síndicos e Administradores podem criar.
    """
    try:
        user = request.user
        is_sindico = user.groups.filter(name="Síndicos").exists()

        if not (user.is_staff or is_sindico):
            return Response(
                {"error": "Apenas Síndicos podem cadastrar unidades."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = UnidadeSerializer(data=request.data)
        if serializer.is_valid():
            unidade = serializer.save(created_by=user)
            return Response(
                UnidadeSerializer(unidade).data,
                status=status.HTTP_201_CREATED,
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response(
            {"error": f"Erro ao criar unidade: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def unidade_create_bulk_view(request):
    """
    Cria múltiplas unidades de uma vez.
    Apenas Síndicos e Administradores podem criar.
    """
    try:
        user = request.user
        is_sindico = user.groups.filter(name="Síndicos").exists()

        if not (user.is_staff or is_sindico):
            return Response(
                {"error": "Apenas Síndicos podem cadastrar unidades."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = UnidadeCreateBulkSerializer(data=request.data)
        if serializer.is_valid():
            unidades = serializer.save()

            # Atualizar created_by de todas as unidades criadas
            for unidade in unidades:
                unidade.created_by = user
                unidade.save()

            return Response(
                {
                    "message": f"{len(unidades)} unidades criadas com sucesso.",
                    "unidades": UnidadeListSerializer(
                        unidades, many=True
                    ).data,
                },
                status=status.HTTP_201_CREATED,
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response(
            {"error": f"Erro ao criar unidades: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def unidade_detail_view(request, pk):
    """
    Obtém detalhes de uma unidade específica
    """
    try:
        # Relação reversa: usar prefetch_related
        unidade = Unidade.objects.prefetch_related("moradores").get(pk=pk)
        serializer = UnidadeSerializer(unidade)
        return Response(serializer.data)

    except Unidade.DoesNotExist:
        return Response(
            {"error": "Unidade não encontrada."},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"error": f"Erro ao obter unidade: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["PUT", "PATCH"])
@permission_classes([IsAuthenticated])
def unidade_update_view(request, pk):
    """
    Atualiza uma unidade.
    Apenas Síndicos e Administradores podem editar.
    """
    try:
        user = request.user
        is_sindico = user.groups.filter(name="Síndicos").exists()

        if not (user.is_staff or is_sindico):
            return Response(
                {"error": "Apenas Síndicos podem editar unidades."},
                status=status.HTTP_403_FORBIDDEN,
            )

        unidade = Unidade.objects.get(pk=pk)
        serializer = UnidadeSerializer(
            unidade, data=request.data, partial=(request.method == "PATCH")
        )

        if serializer.is_valid():
            unidade = serializer.save(updated_by=user)
            return Response(UnidadeSerializer(unidade).data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except Unidade.DoesNotExist:
        return Response(
            {"error": "Unidade não encontrada."},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"error": f"Erro ao atualizar unidade: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def unidade_inactivate_view(request, pk):
    """
    Inativa uma unidade (soft delete).
    Apenas Síndicos e Administradores podem inativar.
    """
    try:
        user = request.user
        is_sindico = user.groups.filter(name="Síndicos").exists()

        if not (user.is_staff or is_sindico):
            return Response(
                {"error": "Apenas Síndicos podem inativar unidades."},
                status=status.HTTP_403_FORBIDDEN,
            )

        unidade = Unidade.objects.get(pk=pk)
        unidade.is_active = False
        unidade.updated_by = user
        unidade.save()

        return Response(
            {
                "message": "Unidade inativada com sucesso.",
                "unidade": UnidadeSerializer(unidade).data,
            }
        )

    except Unidade.DoesNotExist:
        return Response(
            {"error": "Unidade não encontrada."},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"error": f"Erro ao inativar unidade: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def unidade_delete_view(request, pk):
    """
    Exclui uma unidade permanentemente.
    Apenas Administradores podem excluir.
    """
    try:
        user = request.user

        if not user.is_staff:
            return Response(
                {"error": "Apenas administradores podem excluir unidades."},
                status=status.HTTP_403_FORBIDDEN,
            )

        unidade = Unidade.objects.get(pk=pk)
        unidade.delete()

        return Response(
            {"message": "Unidade excluída com sucesso."},
            status=status.HTTP_204_NO_CONTENT,
        )

    except Unidade.DoesNotExist:
        return Response(
            {"error": "Unidade não encontrada."},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"error": f"Erro ao excluir unidade: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# ---------------------------------------------------------------------------
# Colunas do modelo Excel (ordem importa — usada na importação também)
# ---------------------------------------------------------------------------
_EXCEL_COLUMNS = [
    {
        "key": "bloco",
        "label": "Bloco",
        "width": 12,
        "obrigatorio": False,
        "comentario": (
            "Bloco ou torre da unidade.\n"
            "Opcional. Se o condomínio não tiver blocos, deixe em branco.\n"
            "Exemplos: A, B, 1, Torre Norte"
        ),
    },
    {
        "key": "numero_unidade",
        "label": "Número da Unidade",
        "width": 20,
        "obrigatorio": True,
        "comentario": (
            "Número ou identificação da unidade.\n"
            "Obrigatório.\n"
            "Exemplos: 101, 201A, Cobertura"
        ),
    },
]


def _gerar_username(primeiro_nome: str, sobrenome: str) -> str:
    """Replica a lógica de formatUsername do frontend."""
    full = f"{primeiro_nome} {sobrenome}".lower()
    full = unicodedata.normalize("NFD", full)
    full = "".join(c for c in full if unicodedata.category(c) != "Mn")
    full = re.sub(r"[^a-z0-9]", "_", full)
    full = re.sub(r"_+", "_", full)
    return full.strip("_")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_modelo_excel_view(request):
    """
    Retorna um arquivo Excel modelo (.xlsx) formatado para importação de
    unidades. Apenas Síndicos e Administradores podem acessar.
    """
    user = request.user
    is_sindico = user.groups.filter(name="Síndicos").exists()

    if not (user.is_staff or is_sindico):
        return Response(
            {"error": "Apenas Síndicos podem exportar o modelo."},
            status=status.HTTP_403_FORBIDDEN,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidades"

    num_cols = len(_EXCEL_COLUMNS)
    num_data_rows = 500  # linhas reservadas para dados + validação

    # ── Estilos ─────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1a8a72")  # verde escuro
    required_fill = PatternFill(
        "solid", fgColor="2abb98"
    )  # verde principal do sistema
    optional_fill = PatternFill("solid", fgColor="5dd3be")  # verde claro
    example_fill = PatternFill("solid", fgColor="e8f8f5")  # verde pastel
    label_font = Font(bold=True, color="FFFFFF", size=10)
    example_font = Font(color="1a5c4e", italic=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color="1a8a72")
    border_header = Border(left=thick, right=thick, top=thick, bottom=thick)
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Linha 1: Título único ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Modelo de Importação — Unidades"
    title_cell.fill = header_fill
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = center

    # ── Linha 2: Labels amigáveis ────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    for i, col_def in enumerate(_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=2, column=i)
        suffix = " *" if col_def["obrigatorio"] else ""
        cell.value = f"{col_def['label']}{suffix}"
        cell.fill = required_fill if col_def["obrigatorio"] else optional_fill
        cell.font = label_font
        cell.alignment = center
        cell.border = border_header

    # ── Linha 3: Exemplo ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    example_row = ["A", "101"]
    for i, val in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=i)
        cell.value = val
        cell.fill = example_fill
        cell.font = example_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell
        cell.number_format = "@"  # formato texto
        cell.protection = Protection(locked=False)

    # ── Formatar colunas de dados (linhas 4-N) como texto ────────────────────
    for row_idx in range(4, 3 + num_data_rows):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = "@"
            cell.border = border_cell
            cell.protection = Protection(locked=False)

    # ── Largura das colunas ──────────────────────────────────────────────────
    for i, col_def in enumerate(_EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_def["width"]

    # ── Legenda rodapé ───────────────────────────────────────────────────────
    legend_row = 3 + num_data_rows + 1
    ws.merge_cells(
        start_row=legend_row,
        start_column=1,
        end_row=legend_row,
        end_column=num_cols,
    )
    legend_cell = ws.cell(row=legend_row, column=1)
    legend_cell.value = "(*) Campos obrigatórios"
    legend_cell.font = Font(color="595959", italic=True, size=9)
    legend_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Comentários de orientação nos cabeçalhos (linha 2) ───────────────────
    for i, col_def in enumerate(_EXCEL_COLUMNS, start=1):
        comment = Comment(col_def["comentario"], "Sistema")
        comment.width = 280
        comment.height = 120
        ws.cell(row=2, column=i).comment = comment

    # ── Validações de dados (linhas 4 em diante) ─────────────────────────────
    data_start = 4
    data_end = 3 + num_data_rows

    def col_ref(col_key, row):
        idx = next(
            i + 1 for i, c in enumerate(_EXCEL_COLUMNS) if c["key"] == col_key
        )
        return f"{get_column_letter(idx)}{row}"

    def col_range(col_key):
        idx = next(
            i + 1 for i, c in enumerate(_EXCEL_COLUMNS) if c["key"] == col_key
        )
        letter = get_column_letter(idx)
        return f"{letter}{data_start}:{letter}{data_end}"

    # Número da unidade: 1 a 20 caracteres
    dv_unidade = DataValidation(
        type="textLength",
        operator="between",
        formula1="1",
        formula2="20",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Número de unidade inválido",
        error="O número da unidade deve ter entre 1 e 20 caracteres. Ex: 101, 201A",
        showInputMessage=True,
        promptTitle="Número da unidade",
        prompt="Obrigatório. Ex: 101, 201A, Cobertura",
    )
    dv_unidade.sqref = col_range("numero_unidade")
    ws.add_data_validation(dv_unidade)

    # ── Congelar painel abaixo dos cabeçalhos ────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Proteger cabeçalhos (linhas 1 e 2) ───────────────────────────────────
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    # ── Zoom ─────────────────────────────────────────────────────────────────
    ws.sheet_view.zoomScale = 110

    # ── Salvar ───────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="modelo_unidades.xlsx"'
    )
    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_excel_view(request):
    """
    Importa unidades a partir de um arquivo Excel (.xlsx).
    Apenas Síndicos e Administradores podem importar.
    O arquivo deve seguir o modelo gerado por export_modelo_excel_view.
    A linha 2 contém os rótulos das colunas, a linha 3 é o exemplo (ignorada), dados a partir da linha 4.
    Retorna JSON com {'criados': N, 'erros': [{'linha': X, 'motivo': '...'}]}.
    """
    user = request.user
    is_sindico = user.groups.filter(name="Síndicos").exists()

    if not (user.is_staff or is_sindico):
        return Response(
            {"error": "Apenas Síndicos podem importar unidades."},
            status=status.HTTP_403_FORBIDDEN,
        )

    if "arquivo" not in request.FILES:
        return Response(
            {"error": "Nenhum arquivo enviado. Use o campo 'arquivo'."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    arquivo = request.FILES["arquivo"]

    try:
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    except Exception:
        return Response(
            {"error": "Arquivo inválido. Envie um arquivo .xlsx válido."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return Response(
            {
                "error": "Planilha em formato inválido. Use o modelo fornecido pelo sistema."
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Linha 2 (índice 1) contém os rótulos amigáveis das colunas
    label_row_raw = [str(h).strip() if h is not None else "" for h in rows[1]]

    def _norm(s):
        return s.replace(" *", "").strip().lower()

    normalized_labels = [_norm(h) for h in label_row_raw]
    col_positions = {}
    for col_def in _EXCEL_COLUMNS:
        norm = col_def["label"].lower()
        if norm in normalized_labels:
            col_positions[col_def["key"]] = normalized_labels.index(norm)

    expected_keys = [c["key"] for c in _EXCEL_COLUMNS]
    missing_cols = [k for k in expected_keys if k not in col_positions]
    if missing_cols:
        return Response(
            {
                "error": f"Colunas ausentes: {', '.join(missing_cols)}. Use o modelo fornecido."
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Dados começam na linha 4 (índice 3); linhas 1-3 são título, cabeçalho e exemplo
    data_rows_iter = enumerate(rows[3:], start=4)

    def get_col(row, key):
        idx = col_positions.get(key)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    condominio = user.condominio
    criados = 0
    erros = []

    for row_num, row in data_rows_iter:
        # Ignorar linhas completamente vazias
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        bloco = get_col(row, "bloco")
        numero = get_col(row, "numero_unidade")

        # Validação mínima
        if not numero:
            erros.append({"linha": row_num, "motivo": "numero_unidade é obrigatório."})
            continue

        try:
            unidade, created = Unidade.objects.get_or_create(
                numero=numero,
                bloco=bloco if bloco else None,
                defaults={"created_by": user},
            )

            if created:
                criados += 1

        except Exception as e:
            erros.append({"linha": row_num, "motivo": str(e)})

    return Response(
        {"criados": criados, "erros": erros}, status=status.HTTP_200_OK
    )
